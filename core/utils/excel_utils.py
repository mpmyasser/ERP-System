from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

def apply_professional_style(sheet, df):
    """Apply standard professional styling to an Excel sheet"""
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    border_side = Side(style='thin', color='000000')
    full_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 1. Right to Left
    sheet.sheet_view.rightToLeft = True
    
    # 2. Header Style
    for col_idx in range(1, len(df.columns) + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = full_border
        
    # 3. Content Style & Borders
    for r_idx in range(2, len(df) + 2):
        for c_idx in range(1, len(df.columns) + 1):
            cell = sheet.cell(row=r_idx, column=c_idx)
            cell.alignment = center_align
            cell.border = full_border
            
            # Check if this column contains numeric values and apply number format
            col_header = sheet.cell(row=1, column=c_idx).value
            if col_header and any(keyword in str(col_header) for keyword in ['راتب', 'أيام', 'إجمالي', 'حوافز', 'إضافي', 'بدلات', 'تأمينات', 'تأخيرات', 'انصراف', 'جزاءات', 'سلف', 'استقطاعات', 'صافي', 'قيمة', 'خصم', 'مبلغ', 'رصيد']):
                # Apply number format with 2 decimal places
                cell.number_format = '#,##0.00'
            
    # 4. Auto-adjust columns
    for col in sheet.columns:
        max_length = 0
        # Handle merged cells or unexpected types gracefully
        try:
             # Try to get column letter
             if hasattr(col[0], 'column_letter'):
                 column = col[0].column_letter
             else:
                 continue
                 
             for cell in col:
                 try:
                     val_len = len(str(cell.value)) if cell.value else 0
                     if val_len > max_length:
                         max_length = val_len
                 except: pass
                 
             sheet.column_dimensions[column].width = max_length + 3
        except:
            pass
