"""
تطبيق تتبع إنتاج العمال - الإصدار المستقل
Flask + SQLite
"""

import os
import sys
import math
from datetime import datetime, date, time, timedelta

from flask import (
    Flask, render_template, request, jsonify,
    redirect, url_for, flash, send_file
)
from sqlalchemy import func, desc, asc, and_, case

# استيراد الموديلات من نفس المجلد
sys.path.insert(0, os.path.dirname(__file__))
from models import (
    Base, Worker, Stage, Product, ProductionRecord, Benchmark,
    init_and_get_session, DB_PATH
)

app = Flask(__name__)
app.secret_key = 'worker-productivity-secret-key-2025'
app.config['TEMPLATES_AUTO_RELOAD'] = True

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_FULL_PATH = os.path.join(BASE_DIR, DB_PATH)


# ===================== أدوات مساعدة =====================

def get_session():
    return init_and_get_session(DB_FULL_PATH)


def time_to_hours(t: time) -> float:
    """تحويل time object إلى عدد ساعات (float)"""
    return t.hour + t.minute / 60.0 + t.second / 3600.0


def hours_diff(t1: time, t2: time) -> float:
    """الفرق بالساعات بين وقتين"""
    h1 = time_to_hours(t1)
    h2 = time_to_hours(t2)
    diff = h2 - h1
    if diff < 0:
        diff += 24  # عبر منتصف الليل
    return round(diff, 4)


def parse_time(s: str) -> time:
    """تحويل نص إلى time object"""
    s = s.strip()
    for fmt in ('%H:%M:%S', '%H:%M', '%I:%M %p', '%I:%M%p'):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    # محاولة يدوية
    parts = s.split(':')
    if len(parts) >= 2:
        return time(int(parts[0]), int(parts[1]))
    raise ValueError(f"لا يمكن تحويل الوقت: {s}")


def recalc_benchmarks(session):
    """إعادة حساب المعايير الإحصائية لجميع المجموعات"""
    # حذف المعايير القديمة
    session.query(Benchmark).delete()

    # تجميع السجلات حسب (stage_code, product_code)
    from sqlalchemy import text
    rows = session.query(
        ProductionRecord.stage_code,
        ProductionRecord.product_code,
        func.count(ProductionRecord.id).label('cnt')
    ).group_by(
        ProductionRecord.stage_code,
        ProductionRecord.product_code
    ).having(
        func.count(ProductionRecord.id) >= 3  # على الأقل 3 سجلات لحساب إحصائيات ذات معنى
    ).all()

    for row in rows:
        # جلب كل ساعات العمل لكل وحدة لهذه المجموعة
        records = session.query(ProductionRecord).filter(
            ProductionRecord.stage_code == row.stage_code,
            ProductionRecord.product_code == row.product_code,
            ProductionRecord.quantity > 0,
            ProductionRecord.hours_worked > 0
        ).all()

        if len(records) < 3:
            continue

        units = []
        for r in records:
            hp = r.hours_worked / r.quantity
            units.append(hp)

        units_sorted = sorted(units)
        n = len(units_sorted)
        avg_val = sum(units) / n

        # الوسيط
        if n % 2 == 1:
            med_val = units_sorted[n // 2]
        else:
            med_val = (units_sorted[n // 2 - 1] + units_sorted[n // 2]) / 2

        # الانحراف المعياري
        variance = sum((x - avg_val) ** 2 for x in units) / n
        std_val = math.sqrt(variance)

        # Percentiles
        def percentile(data, p):
            k = (p / 100.0) * (len(data) - 1)
            f = math.floor(k)
            c = math.ceil(k)
            if f == c:
                return data[int(k)]
            d0 = data[int(f)] * (c - k)
            d1 = data[int(c)] * (k - f)
            return d0 + d1

        b = Benchmark(
            stage_code=row.stage_code,
            product_code=row.product_code,
            avg_hours_per_unit=round(avg_val, 6),
            median_hours_per_unit=round(med_val, 6),
            min_hours_per_unit=round(units_sorted[0], 6),
            max_hours_per_unit=round(units_sorted[-1], 6),
            std_hours_per_unit=round(std_val, 6),
            p25_hours_per_unit=round(percentile(units_sorted, 25), 6),
            p75_hours_per_unit=round(percentile(units_sorted, 75), 6),
            p90_hours_per_unit=round(percentile(units_sorted, 90), 6),
            record_count=n,
            last_updated=datetime.now()
        )
        session.add(b)

    session.commit()


# ===================== الصفحات الرئيسية =====================

@app.route('/')
def index():
    """الصفحة الرئيسية - Dashboard"""
    session = get_session()
    try:
        # إحصائيات سريعة
        worker_count = session.query(func.count(Worker.id)).filter(Worker.is_active == 1).scalar() or 0
        total_records = session.query(func.count(ProductionRecord.id)).scalar() or 0
        total_quantity = session.query(func.sum(ProductionRecord.quantity)).scalar() or 0

        # آخر 10 سجلات
        recent_records = session.query(ProductionRecord).order_by(
            ProductionRecord.created_at.desc()
        ).limit(10).all()

        return render_template(
            'index.html',
            worker_count=worker_count,
            total_records=total_records,
            total_quantity=round(total_quantity, 0),
            recent_records=recent_records
        )
    finally:
        session.close()


@app.route('/entry')
def entry_page():
    """صفحة إدخال بيانات الإنتاج"""
    session = get_session()
    try:
        workers = session.query(Worker).filter(Worker.is_active == 1).order_by(Worker.name).all()
        stages = session.query(Stage).order_by(Stage.code).all()
        products = session.query(Product).order_by(Product.name, Product.size).all()
        return render_template(
            'entry.html',
            workers=workers,
            stages=stages,
            products=products,
            today=date.today().isoformat()
        )
    finally:
        session.close()


@app.route('/reports')
def reports_page():
    """صفحة التقارير والمقارنات"""
    return render_template('reports.html')


# ===================== API - إدارة البيانات الأساسية =====================

@app.route('/api/workers', methods=['GET'])
def api_workers_list():
    session = get_session()
    try:
        workers = session.query(Worker).order_by(Worker.code).all()
        result = []
        for w in workers:
            record_count = session.query(func.count(ProductionRecord.id)).filter(
                ProductionRecord.worker_code == w.code
            ).scalar() or 0
            result.append({
                'id': w.id,
                'code': w.code,
                'name': w.name,
                'hire_date': w.hire_date.isoformat() if w.hire_date else '',
                'is_insured': w.is_insured,
                'salary': w.salary,
                'is_active': w.is_active,
                'record_count': record_count
            })
        return jsonify({'ok': True, 'workers': result})
    finally:
        session.close()


@app.route('/api/workers/add', methods=['POST'])
def api_workers_add():
    data = request.get_json() or {}
    session = get_session()
    try:
        code = str(data.get('code', '')).strip()
        name = str(data.get('name', '')).strip()
        if not code or not name:
            return jsonify({'ok': False, 'message': 'الكود والاسم مطلوبان'}), 400

        existing = session.query(Worker).filter(Worker.code == code).first()
        if existing:
            return jsonify({'ok': False, 'message': f'العامل بالكود {code} موجود بالفعل'}), 400

        hire = None
        if data.get('hire_date'):
            try:
                hire = datetime.strptime(data['hire_date'], '%Y-%m-%d').date()
            except ValueError:
                pass

        worker = Worker(
            code=code,
            name=name,
            hire_date=hire,
            is_insured=str(data.get('is_insured', 'غير مؤمن')),
            salary=float(data.get('salary', 0)),
            is_active=1
        )
        session.add(worker)
        session.commit()
        return jsonify({'ok': True, 'worker': {'code': worker.code, 'name': worker.name}})
    except Exception as e:
        session.rollback()
        return jsonify({'ok': False, 'message': str(e)}), 500
    finally:
        session.close()


@app.route('/api/workers/import', methods=['POST'])
def api_workers_import():
    """استيراد عمال من JSON batch"""
    data = request.get_json() or {}
    items = data.get('items', [])
    session = get_session()
    try:
        added = 0
        for item in items:
            code = str(item.get('code', '')).strip()
            name = str(item.get('name', '')).strip()
            if not code or not name:
                continue
            existing = session.query(Worker).filter(Worker.code == code).first()
            if existing:
                continue
            worker = Worker(
                code=code,
                name=name,
                hire_date=datetime.strptime(item['hire_date'], '%Y-%m-%d').date() if item.get('hire_date') else None,
                is_insured=str(item.get('is_insured', 'غير مؤمن')),
                salary=float(item.get('salary', 0)),
                is_active=1
            )
            session.add(worker)
            added += 1
        session.commit()
        return jsonify({'ok': True, 'added': added})
    except Exception as e:
        session.rollback()
        return jsonify({'ok': False, 'message': str(e)}), 500
    finally:
        session.close()


@app.route('/api/workers/delete', methods=['POST'])
def api_workers_delete():
    data = request.get_json() or {}
    session = get_session()
    try:
        code = str(data.get('code', '')).strip()
        worker = session.query(Worker).filter(Worker.code == code).first()
        if not worker:
            return jsonify({'ok': False, 'message': 'العامل غير موجود'}), 404
        # حذف السجلات المرتبطة
        session.query(ProductionRecord).filter(ProductionRecord.worker_code == code).delete()
        session.delete(worker)
        session.commit()
        return jsonify({'ok': True})
    except Exception as e:
        session.rollback()
        return jsonify({'ok': False, 'message': str(e)}), 500
    finally:
        session.close()


@app.route('/api/stages', methods=['GET'])
def api_stages_list():
    session = get_session()
    try:
        stages = session.query(Stage).order_by(Stage.code).all()
        return jsonify({
            'ok': True,
            'stages': [{
                'code': s.code,
                'name': s.name,
                'machine_type': s.machine_type or '',
                'product_type': s.product_type or ''
            } for s in stages]
        })
    finally:
        session.close()


@app.route('/api/stages/add', methods=['POST'])
def api_stages_add():
    data = request.get_json() or {}
    session = get_session()
    try:
        code = str(data.get('code', '')).strip()
        name = str(data.get('name', '')).strip()
        if not code or not name:
            return jsonify({'ok': False, 'message': 'كود واسم المرحلة مطلوبان'}), 400
        existing = session.query(Stage).filter(Stage.code == code).first()
        if existing:
            return jsonify({'ok': False, 'message': f'المرحلة {code} موجودة بالفعل'}), 400
        stage = Stage(
            code=code,
            name=name,
            machine_type=str(data.get('machine_type', '')),
            product_type=str(data.get('product_type', ''))
        )
        session.add(stage)
        session.commit()
        return jsonify({'ok': True, 'stage': {'code': stage.code, 'name': stage.name}})
    except Exception as e:
        session.rollback()
        return jsonify({'ok': False, 'message': str(e)}), 500
    finally:
        session.close()


@app.route('/api/products', methods=['GET'])
def api_products_list():
    session = get_session()
    try:
        products = session.query(Product).order_by(Product.name, Product.size).all()
        return jsonify({
            'ok': True,
            'products': [{
                'code': p.code,
                'name': p.name,
                'size': p.size
            } for p in products]
        })
    finally:
        session.close()


@app.route('/api/products/add', methods=['POST'])
def api_products_add():
    data = request.get_json() or {}
    session = get_session()
    try:
        code = str(data.get('code', '')).strip()
        name = str(data.get('name', '')).strip()
        size = str(data.get('size', '')).strip()
        if not code or not name or not size:
            return jsonify({'ok': False, 'message': 'الكود والاسم والمقاس مطلوبون'}), 400
        existing = session.query(Product).filter(
            Product.code == code,
            Product.size == size
        ).first()
        if existing:
            return jsonify({'ok': False, 'message': f'الصنف {code} بالمقاس {size} موجود بالفعل'}), 400
        prod = Product(code=code, name=name, size=size)
        session.add(prod)
        session.commit()
        return jsonify({'ok': True, 'product': {'code': prod.code, 'name': prod.name, 'size': prod.size}})
    except Exception as e:
        session.rollback()
        return jsonify({'ok': False, 'message': str(e)}), 500
    finally:
        session.close()


# ===================== API - سجلات الإنتاج =====================

@app.route('/api/records/add', methods=['POST'])
def api_records_add():
    """
    إضافة سجلات إنتاج (يدعم متعدد في مرة واحدة)
    Body: {"records": [...], "date": "2025-01-01"}
    """
    data = request.get_json() or {}
    session = get_session()
    try:
        items = data.get('records', [])
        if not items:
            return jsonify({'ok': False, 'message': 'لا توجد سجلات'}), 400

        default_date = data.get('date', '')
        added = 0
        errors = []

        for i, item in enumerate(items):
            try:
                # تاريخ السجل
                record_date_str = item.get('record_date', default_date)
                if not record_date_str:
                    errors.append(f"الصف {i+1}: التاريخ مطلوب")
                    continue

                try:
                    record_date = datetime.strptime(record_date_str, '%Y-%m-%d').date()
                except ValueError:
                    errors.append(f"الصف {i+1}: تاريخ غير صالح '{record_date_str}'")
                    continue

                worker_code = str(item.get('worker_code', '')).strip()
                stage_code = str(item.get('stage_code', '')).strip()
                product_code = str(item.get('product_code', '')).strip()
                quantity_str = str(item.get('quantity', '')).strip()
                time_from_str = str(item.get('time_from', '')).strip()
                time_to_str = str(item.get('time_to', '')).strip()

                if not worker_code:
                    errors.append(f"الصف {i+1}: كود العامل مطلوب")
                    continue
                if not stage_code:
                    errors.append(f"الصف {i+1}: كود المرحلة مطلوب")
                    continue
                if not product_code:
                    errors.append(f"الصف {i+1}: كود الصنف مطلوب")
                    continue

                quantity = float(quantity_str) if quantity_str else 0
                if quantity <= 0:
                    errors.append(f"الصف {i+1}: الكمية يجب أن تكون أكبر من صفر")
                    continue

                # تحويل الأوقات
                try:
                    tf = parse_time(time_from_str) if time_from_str else time(8, 0)
                    tt = parse_time(time_to_str) if time_to_str else time(16, 0)
                except ValueError as e:
                    errors.append(f"الصف {i+1}: {str(e)}")
                    continue

                # حساب ساعات العمل
                hw = hours_diff(tf, tt)
                if hw <= 0:
                    errors.append(f"الصف {i+1}: وقت الانتهاء يجب أن يكون بعد وقت البدء")
                    continue

                # التحقق من وجود العامل والمرحلة والصنف
                worker = session.query(Worker).filter(Worker.code == worker_code).first()
                if not worker:
                    errors.append(f"الصف {i+1}: العامل '{worker_code}' غير موجود")
                    continue

                stage = session.query(Stage).filter(Stage.code == stage_code).first()
                if not stage:
                    errors.append(f"الصف {i+1}: المرحلة '{stage_code}' غير موجودة")
                    continue

                prod = session.query(Product).filter(Product.code == product_code).first()
                if not prod:
                    errors.append(f"الصف {i+1}: الصنف '{product_code}' غير موجود")
                    continue

                record = ProductionRecord(
                    record_date=record_date,
                    worker_code=worker_code,
                    stage_code=stage_code,
                    product_code=product_code,
                    time_from=tf,
                    time_to=tt,
                    hours_worked=hw,
                    quantity=quantity,
                    machine_type=item.get('machine_type', ''),
                    notes=item.get('notes', '')
                )
                session.add(record)
                added += 1

            except Exception as e:
                errors.append(f"الصف {i+1}: {str(e)}")

        if added > 0:
            session.commit()
            # إعادة حساب المعايير بعد الإضافة
            recalc_benchmarks(session)

        return jsonify({
            'ok': True,
            'added': added,
            'errors': errors
        })

    except Exception as e:
        session.rollback()
        return jsonify({'ok': False, 'message': str(e)}), 500
    finally:
        session.close()


@app.route('/api/records/list')
def api_records_list():
    """قائمة سجلات الإنتاج مع فلترة"""
    session = get_session()
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        worker_code = request.args.get('worker_code', '').strip()
        stage_code = request.args.get('stage_code', '').strip()
        date_from = request.args.get('date_from', '').strip()
        date_to = request.args.get('date_to', '').strip()

        query = session.query(ProductionRecord)

        if worker_code:
            query = query.filter(ProductionRecord.worker_code == worker_code)
        if stage_code:
            query = query.filter(ProductionRecord.stage_code == stage_code)
        if date_from:
            query = query.filter(ProductionRecord.record_date >= datetime.strptime(date_from, '%Y-%m-%d').date())
        if date_to:
            query = query.filter(ProductionRecord.record_date <= datetime.strptime(date_to, '%Y-%m-%d').date())

        total = query.count()
        records = query.order_by(ProductionRecord.record_date.desc(), ProductionRecord.id.desc()).offset(
            (page - 1) * per_page).limit(per_page).all()

        result = []
        for r in records:
            result.append({
                'id': r.id,
                'record_date': r.record_date.isoformat(),
                'worker_code': r.worker_code,
                'stage_code': r.stage_code,
                'product_code': r.product_code,
                'time_from': r.time_from.strftime('%H:%M'),
                'time_to': r.time_to.strftime('%H:%M'),
                'hours_worked': r.hours_worked,
                'quantity': r.quantity,
                'machine_type': r.machine_type or '',
                'notes': r.notes or ''
            })

        return jsonify({
            'ok': True,
            'records': result,
            'total': total,
            'page': page,
            'per_page': per_page,
            'total_pages': math.ceil(total / per_page) if total > 0 else 1
        })
    finally:
        session.close()


@app.route('/api/records/delete', methods=['POST'])
def api_records_delete():
    """حذف سجل إنتاج"""
    data = request.get_json() or {}
    session = get_session()
    try:
        record_id = int(data.get('id', 0))
        if not record_id:
            return jsonify({'ok': False, 'message': 'رقم السجل مطلوب'}), 400
        record = session.query(ProductionRecord).filter(ProductionRecord.id == record_id).first()
        if not record:
            return jsonify({'ok': False, 'message': 'السجل غير موجود'}), 404
        session.delete(record)
        session.commit()
        recalc_benchmarks(session)
        return jsonify({'ok': True})
    except Exception as e:
        session.rollback()
        return jsonify({'ok': False, 'message': str(e)}), 500
    finally:
        session.close()


# ===================== API - التقارير والتحليلات =====================

@app.route('/api/reports/worker-ranking')
def api_reports_worker_ranking():
    """
    ترتيب العمال حسب السرعة والإنتاجية
    Query params: sort_by (speed / productivity / quantity), period (all / month / week)
    """
    session = get_session()
    try:
        sort_by = request.args.get('sort_by', 'speed')
        limit = request.args.get('limit', 20, type=int)

        # فلترة حسب الفترة
        date_filter = None
        period = request.args.get('period', 'all')
        if period == 'month':
            date_filter = date.today().replace(day=1)
        elif period == 'week':
            today = date.today()
            date_filter = today - timedelta(days=today.weekday())

        # جلب جميع العمال النشطين
        workers = session.query(Worker).filter(Worker.is_active == 1).all()
        ranking = []

        for w in workers:
            query = session.query(ProductionRecord).filter(
                ProductionRecord.worker_code == w.code,
                ProductionRecord.quantity > 0,
                ProductionRecord.hours_worked > 0
            )
            if date_filter:
                query = query.filter(ProductionRecord.record_date >= date_filter)

            records = query.all()
            if not records:
                continue

            total_qty = sum(r.quantity for r in records)
            total_hours = sum(r.hours_worked for r in records)
            overall_speed = total_qty / total_hours if total_hours > 0 else 0

            # حساب سرعة العامل مقارنة بالمعايير
            speed_ratio_sum = 0
            speed_count = 0
            for r in records:
                b = session.query(Benchmark).filter(
                    Benchmark.stage_code == r.stage_code,
                    Benchmark.product_code == r.product_code
                ).first()
                if b and b.avg_hours_per_unit and b.avg_hours_per_unit > 0:
                    # الوقت القياسي للوحدة / وقت العامل للوحدة
                    worker_unit_time = r.hours_worked / r.quantity if r.quantity > 0 else 0
                    if worker_unit_time > 0:
                        ratio = (b.avg_hours_per_unit / worker_unit_time) * 100
                        speed_ratio_sum += ratio
                        speed_count += 1

            avg_speed_ratio = round(speed_ratio_sum / speed_count, 1) if speed_count > 0 else 0

            ranking.append({
                'worker_code': w.code,
                'worker_name': w.name,
                'total_quantity': round(total_qty, 0),
                'total_hours': round(total_hours, 2),
                'overall_speed': round(overall_speed, 2),  # وحدة/ساعة
                'speed_ratio': avg_speed_ratio,             # نسبة مئوية
                'record_count': len(records)
            })

        # ترتيب حسب الاختيار
        if sort_by == 'speed':
            ranking.sort(key=lambda x: x['speed_ratio'], reverse=True)
        elif sort_by == 'productivity':
            ranking.sort(key=lambda x: x['overall_speed'], reverse=True)
        else:  # quantity
            ranking.sort(key=lambda x: x['total_quantity'], reverse=True)

        # إضافة الترتيب
        for i, r in enumerate(ranking):
            r['rank'] = i + 1

        return jsonify({'ok': True, 'ranking': ranking[:limit], 'total': len(ranking)})
    finally:
        session.close()


@app.route('/api/reports/stage-analysis')
def api_reports_stage_analysis():
    """تحليل المراحل - أفضل عامل في كل مرحلة"""
    session = get_session()
    try:
        stages = session.query(Stage).order_by(Stage.code).all()
        result = []

        for st in stages:
            # لكل مرحلة، جلب جميع السجلات
            records = session.query(ProductionRecord).filter(
                ProductionRecord.stage_code == st.code,
                ProductionRecord.quantity > 0,
                ProductionRecord.hours_worked > 0
            ).all()

            if not records:
                continue

            # تجميع حسب العامل
            worker_stats = {}
            for r in records:
                if r.worker_code not in worker_stats:
                    worker_stats[r.worker_code] = {
                        'total_qty': 0,
                        'total_hours': 0,
                        'total_time': 0
                    }
                worker_stats[r.worker_code]['total_qty'] += r.quantity
                worker_stats[r.worker_code]['total_hours'] += r.hours_worked
                worker_stats[r.worker_code]['total_time'] += r.hours_worked / r.quantity if r.quantity > 0 else 0

            # تحويل إلى قائمة وترتيب حسب السرعة
            stage_workers = []
            for wc, ws in worker_stats.items():
                worker = session.query(Worker).filter(Worker.code == wc).first()
                speed = ws['total_qty'] / ws['total_hours'] if ws['total_hours'] > 0 else 0
                stage_workers.append({
                    'worker_code': wc,
                    'worker_name': worker.name if worker else wc,
                    'total_quantity': round(ws['total_qty'], 0),
                    'total_hours': round(ws['total_hours'], 2),
                    'speed': round(speed, 2)
                })

            stage_workers.sort(key=lambda x: x['speed'], reverse=True)

            # المعايير
            benchmarks = session.query(Benchmark).filter(
                Benchmark.stage_code == st.code
            ).all()
            avg_benchmark = 0
            if benchmarks:
                vals = [b.avg_hours_per_unit for b in benchmarks if b.avg_hours_per_unit]
                avg_benchmark = round(sum(vals) / len(vals), 6) if vals else 0

            result.append({
                'stage_code': st.code,
                'stage_name': st.name,
                'machine_type': st.machine_type or '',
                'worker_count': len(stage_workers),
                'total_records': len(records),
                'workers': stage_workers[:5],  # أفضل 5
                'avg_benchmark_unit_time': avg_benchmark
            })

        return jsonify({'ok': True, 'stages': result})
    finally:
        session.close()


@app.route('/api/reports/dashboard')
def api_reports_dashboard():
    """إحصائيات سريعة للـ Dashboard"""
    session = get_session()
    try:
        today = date.today()
        first_of_month = today.replace(day=1)

        # إجمالي
        total_workers = session.query(func.count(Worker.id)).filter(Worker.is_active == 1).scalar() or 0
        total_records = session.query(func.count(ProductionRecord.id)).scalar() or 0
        total_qty = session.query(func.sum(ProductionRecord.quantity)).scalar() or 0

        # هذا الشهر
        month_records = session.query(func.count(ProductionRecord.id)).filter(
            ProductionRecord.record_date >= first_of_month
        ).scalar() or 0
        month_qty = session.query(func.sum(ProductionRecord.quantity)).filter(
            ProductionRecord.record_date >= first_of_month
        ).scalar() or 0

        # اليوم
        today_records = session.query(func.count(ProductionRecord.id)).filter(
            ProductionRecord.record_date == today
        ).scalar() or 0
        today_qty = session.query(func.sum(ProductionRecord.quantity)).filter(
            ProductionRecord.record_date == today
        ).scalar() or 0
        today_workers = session.query(func.count(
            func.distinct(ProductionRecord.worker_code)
        )).filter(
            ProductionRecord.record_date == today
        ).scalar() or 0

        return jsonify({
            'ok': True,
            'total_workers': total_workers,
            'total_records': total_records,
            'total_quantity': round(total_qty, 0),
            'month_records': month_records,
            'month_quantity': round(month_qty, 0),
            'today_records': today_records,
            'today_quantity': round(today_qty, 0),
            'today_workers': today_workers
        })
    finally:
        session.close()


@app.route('/api/import-excel-data', methods=['POST'])
def api_import_excel_data():
    """استيراد البيانات من ملف Excel الموجود (الانتاج - 2025-سمر.xlsm)"""
    import openpyxl
    from datetime import timedelta

    data = request.get_json() or {}
    file_path = str(data.get('file_path', '')).strip()
    if not file_path:
        return jsonify({'ok': False, 'message': 'مسار الملف مطلوب'}), 400

    if not os.path.exists(file_path):
        return jsonify({'ok': False, 'message': f'الملف غير موجود: {file_path}'}), 400

    session = get_session()
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        errors = []
        workers_added = 0
        stages_added = 0
        products_added = 0
        records_added = 0

        # ----- 1. استيراد العمال من شيت "all" -----
        if 'all' in wb.sheetnames:
            ws = wb['all']
            for row in ws.iter_rows(min_row=2, values_only=True):
                code = str(row[0]).strip() if row[0] else ''
                name = str(row[1]).strip() if row[1] else ''
                if not code or not name:
                    continue
                try:
                    existing = session.query(Worker).filter(Worker.code == code).first()
                    if not existing:
                        hire_str = str(row[2]).strip() if row[2] else ''
                        hire_date = None
                        if hire_str:
                            try:
                                if '-' in hire_str:
                                    hire_date = datetime.strptime(hire_str[:10], '%Y-%m-%d').date()
                                else:
                                    dt = datetime.fromisoformat(hire_str)
                                    hire_date = dt.date() if hasattr(dt, 'date') else dt
                            except (ValueError, TypeError):
                                pass
                        worker = Worker(
                            code=code,
                            name=name,
                            hire_date=hire_date,
                            is_insured=str(row[3] if row[3] else 'غير مؤمن'),
                            salary=float(row[4]) if row[4] else 0,
                            is_active=1
                        )
                        session.add(worker)
                        workers_added += 1
                except Exception as e:
                    errors.append(f"عامل {code}: {str(e)}")

        # ----- 2. استيراد المراحل من شيت "المراحل" -----
        if 'المراحل' in wb.sheetnames:
            ws = wb['المراحل']
            for row in ws.iter_rows(min_row=2, values_only=True):
                code = str(row[0]).strip() if row[0] else ''
                name = str(row[1]).strip() if row[1] else ''
                if not code or not name:
                    continue
                try:
                    existing = session.query(Stage).filter(Stage.code == code).first()
                    if not existing:
                        stage = Stage(
                            code=code,
                            name=name,
                            machine_type=str(row[2] if row[2] else ''),
                            product_type=str(row[3] if row[3] else '')
                        )
                        session.add(stage)
                        stages_added += 1
                except Exception as e:
                    errors.append(f"مرحلة {code}: {str(e)}")

        # ----- 3. استيراد الأصناف من شيت "الاصناف" -----
        if 'الاصناف' in wb.sheetnames:
            ws = wb['الاصناف']
            for row in ws.iter_rows(min_row=2, values_only=True):
                code = str(row[0]).strip() if row[0] else ''
                name = str(row[1]).strip() if row[1] else ''
                size = str(row[2]).strip() if row[2] else ''
                if not code or not name or not size:
                    continue
                try:
                    existing = session.query(Product).filter(
                        Product.code == code,
                        Product.size == size
                    ).first()
                    if not existing:
                        prod = Product(code=code, name=name, size=size)
                        session.add(prod)
                        products_added += 1
                except Exception as e:
                    errors.append(f"صنف {code}: {str(e)}")

        session.commit()

        # ----- 4. استيراد سجلات الإنتاج من شيت "قاعدة البيانات" -----
        if 'قاعدة البيانات' in wb.sheetnames:
            ws = wb['قاعدة البيانات']
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    # التحقق من وجود بيانات كافية
                    if not row[0] or not row[4] or not row[5] or not row[13]:
                        continue

                    # التاريخ
                    date_val = row[0]
                    if isinstance(date_val, datetime):
                        record_date = date_val.date()
                    elif isinstance(date_val, date):
                        record_date = date_val
                    else:
                        continue

                    # العامل
                    worker_code = str(row[5]).strip() if row[5] else ''
                    stage_code = str(row[7]).strip() if row[7] else ''
                    product_code = str(row[10]).strip() if row[10] else ''
                    quantity = float(row[13]) if row[13] else 0

                    if not worker_code or not stage_code or not product_code or quantity <= 0:
                        continue

                    # الأوقات
                    from_val = row[1]
                    to_val = row[2]
                    hours_val = row[3]

                    if isinstance(from_val, datetime):
                        time_from = from_val.time()
                    elif isinstance(from_val, timedelta):
                        total_seconds = int(from_val.total_seconds())
                        h = total_seconds // 3600
                        m = (total_seconds % 3600) // 60
                        s = total_seconds % 60
                        time_from = time(h % 24, m, s)
                    elif isinstance(from_val, time):
                        time_from = from_val
                    else:
                        time_from = time(8, 0)

                    if isinstance(to_val, datetime):
                        time_to = to_val.time()
                    elif isinstance(to_val, timedelta):
                        total_seconds = int(to_val.total_seconds())
                        h = total_seconds // 3600
                        m = (total_seconds % 3600) // 60
                        s = total_seconds % 60
                        time_to = time(h % 24, m, s)
                    elif isinstance(to_val, time):
                        time_to = to_val
                    else:
                        time_to = time(16, 0)

                    # ساعات العمل
                    if isinstance(hours_val, (int, float)):
                        hours_worked = float(hours_val)
                    elif isinstance(hours_val, timedelta):
                        hours_worked = hours_val.total_seconds() / 3600
                    elif isinstance(hours_val, time):
                        hours_worked = hours_val.hour + hours_val.minute / 60
                    else:
                        hours_worked = hours_diff(time_from, time_to)

                    machine_type = str(row[14]).strip() if len(row) > 14 and row[14] else ''

                    # التحقق من وجود العامل والمرحلة والصنف
                    worker = session.query(Worker).filter(Worker.code == worker_code).first()
                    stage = session.query(Stage).filter(Stage.code == stage_code).first()
                    prod = session.query(Product).filter(Product.code == product_code).first()

                    if not worker or not stage or not prod:
                        continue

                    record = ProductionRecord(
                        record_date=record_date,
                        worker_code=worker_code,
                        stage_code=stage_code,
                        product_code=product_code,
                        time_from=time_from,
                        time_to=time_to,
                        hours_worked=hours_worked,
                        quantity=quantity,
                        machine_type=machine_type,
                    )
                    session.add(record)
                    records_added += 1

                except Exception as e:
                    errors.append(f"سجل: {str(e)}")

            session.commit()

        # إعادة حساب المعايير
        if records_added > 0:
            recalc_benchmarks(session)

        wb.close()
        return jsonify({
            'ok': True,
            'workers': workers_added,
            'stages': stages_added,
            'products': products_added,
            'records': records_added,
            'errors': errors[:20]
        })

    except Exception as e:
        session.rollback()
        return jsonify({'ok': False, 'message': str(e)}), 500
    finally:
        session.close()


@app.route('/api/reports/worker-detail')
def api_reports_worker_detail():
    """تفاصيل عامل معين - تحليل أدائه"""
    session = get_session()
    try:
        worker_code = request.args.get('worker_code', '').strip()
        if not worker_code:
            return jsonify({'ok': False, 'message': 'كود العامل مطلوب'}), 400

        worker = session.query(Worker).filter(Worker.code == worker_code).first()
        if not worker:
            return jsonify({'ok': False, 'message': 'العامل غير موجود'}), 404

        records = session.query(ProductionRecord).filter(
            ProductionRecord.worker_code == worker_code,
            ProductionRecord.quantity > 0,
            ProductionRecord.hours_worked > 0
        ).order_by(ProductionRecord.record_date.desc()).limit(500).all()

        total_qty = sum(r.quantity for r in records)
        total_hours = sum(r.hours_worked for r in records)
        overall_speed = total_qty / total_hours if total_hours > 0 else 0

        # تحليل حسب المرحلة
        stage_analysis = {}
        for r in records:
            if r.stage_code not in stage_analysis:
                stage_obj = session.query(Stage).filter(Stage.code == r.stage_code).first()
                stage_analysis[r.stage_code] = {
                    'stage_name': stage_obj.name if stage_obj else r.stage_code,
                    'total_qty': 0,
                    'total_hours': 0
                }
            stage_analysis[r.stage_code]['total_qty'] += r.quantity
            stage_analysis[r.stage_code]['total_hours'] += r.hours_worked

        stages_detail = []
        for sc, sa in stage_analysis.items():
            speed = sa['total_qty'] / sa['total_hours'] if sa['total_hours'] > 0 else 0

            # مقارنة بالمعيار
            benchmarks = session.query(Benchmark).filter(
                Benchmark.stage_code == sc
            ).all()
            avg_bench = 0
            if benchmarks:
                vals = [b.avg_hours_per_unit for b in benchmarks if b.avg_hours_per_unit]
                avg_bench = sum(vals) / len(vals) if vals else 0

            # حساب النسبة
            worker_unit = sa['total_hours'] / sa['total_qty'] if sa['total_qty'] > 0 else 0
            ratio = (avg_bench / worker_unit) * 100 if (avg_bench > 0 and worker_unit > 0) else 0

            stages_detail.append({
                'stage_code': sc,
                'stage_name': sa['stage_name'],
                'total_quantity': round(sa['total_qty'], 0),
                'total_hours': round(sa['total_hours'], 2),
                'speed': round(speed, 2),
                'benchmark_unit_time': round(avg_bench, 6),
                'speed_ratio': round(ratio, 1)
            })

        stages_detail.sort(key=lambda x: x['speed'], reverse=True)

        return jsonify({
            'ok': True,
            'worker': {
                'code': worker.code,
                'name': worker.name,
                'hire_date': worker.hire_date.isoformat() if worker.hire_date else '',
                'is_insured': worker.is_insured,
                'salary': worker.salary,
                'total_quantity': round(total_qty, 0),
                'total_hours': round(total_hours, 2),
                'overall_speed': round(overall_speed, 2),
                'record_count': len(records)
            },
            'stage_analysis': stages_detail
        })
    finally:
        session.close()


# ===================== تشغيل التطبيق =====================

if __name__ == '__main__':
    # تهيئة قاعدة البيانات
    print(f"🗄️  قاعدة البيانات: {DB_FULL_PATH}")
    session = init_and_get_session(DB_FULL_PATH)
    session.close()
    print(f"✅ تم تهيئة قاعدة البيانات بنجاح")
    print(f"🚀 تشغيل التطبيق على http://127.0.0.1:5003")
    app.run(debug=True, host='127.0.0.1', port=5003)