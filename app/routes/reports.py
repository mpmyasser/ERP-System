"""
Reports Routes
==============
Various HR reports
"""

from flask import Blueprint, render_template, request, flash, redirect, url_for, current_app, jsonify
import sys
import os
from datetime import datetime, date, timedelta
from utils.helpers import parse_date_compact, format_date_ar
from app.routes.auth import login_required, permission_required, safe_referrer
import pandas as pd
import io
from flask import send_file
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from core.utils.excel_utils import apply_professional_style

sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'core'))


# Function apply_professional_style moved to core/utils/excel_utils.py

reports_bp = Blueprint('reports', __name__)

@reports_bp.route('/')
def index():
    """Reports main page"""
    return render_template('reports/index.html')

@reports_bp.route('/insurance_costs')
def insurance_costs():
    """Report on insurance costs per department"""
    db = current_app.db
    session = db.get_session()
    
    try:
        from sqlalchemy.orm import load_only
        # Get all departments with only necessary fields
        departments = session.query(Department).options(load_only(Department.id, Department.name)).all()
        # Get all active insured employees with only fields needed for insurance calculation
        employees = session.query(Employee).options(
            load_only(
                Employee.id, Employee.name, Employee.department_id, 
                Employee.is_active, Employee.is_insured, Employee.insurance_salary,
                Employee.insurance_employee_share, Employee.insurance_company_share,
                Employee.insurance_policy, Employee.basic_salary
            )
        ).filter_by(is_active=True, is_insured=True).all()
        
        # Structure data by department
        report_data = []
        grand_total_emp = 0
        grand_total_comp = 0
        grand_total_ins = 0
        
        for dept in departments:
            dept_emps = [e for e in employees if e.department_id == dept.id]
            if not dept_emps:
                continue
                
            dept_stats = {
                'department_name': dept.name,
                'count': len(dept_emps),
                'employee_deductions': 0,
                'company_costs': 0,
                'total_insurance': 0
            }
            
            for emp in dept_emps:
                # Use our new calculation method
                vals = emp.calculate_insurance_values()
                dept_stats['employee_deductions'] += vals['employee_deduction']
                dept_stats['company_costs'] += vals['company_cost']
                dept_stats['total_insurance'] += vals['total_insurance']
            
            report_data.append(dept_stats)
            grand_total_emp += dept_stats['employee_deductions']
            grand_total_comp += dept_stats['company_costs']
            grand_total_ins += dept_stats['total_insurance']
            
        return render_template('reports/insurance_costs.html',
                             report_data=report_data,
                             grand_total_emp=grand_total_emp,
                             grand_total_comp=grand_total_comp,
                             grand_total_ins=grand_total_ins)
    finally:
        session.close()

@reports_bp.route('/insured_employees')
def insured_employees():
    """Detailed report for all insured employees with complete insurance data"""
    db = current_app.db
    session = db.get_session()
    
    try:
        # Get filter parameters
        department_id = request.args.get('department_id', type=int)
        search_query = request.args.get('search', '').strip()
        
        # Get all insured active employees
        employees = session.query(Employee).filter_by(is_active=True, is_insured=True).all()
        
        # Filter by department if specified
        if department_id:
            employees = [e for e in employees if e.department_id == department_id]
        
        # Apply search filter
        if search_query:
            search_lower = search_query.lower()
            employees = [e for e in employees if 
                        (e.name and search_lower in e.name.lower()) or
                        (e.code and search_lower in e.code.lower()) or
                        (e.insurance_number and search_lower in e.insurance_number.lower())]
        
        # Prepare detailed data for each employee
        report_data = []
        total_employee_deduction = 0
        total_company_cost = 0
        total_insurance = 0
        
        for emp in employees:
            insurance_vals = emp.calculate_insurance_values()
            
            # Format insurance start and end dates
            start_date_str = emp.insurance_start_date.strftime('%d/%m/%Y') if emp.insurance_start_date else 'غير محدد'
            end_date_str = emp.insurance_end_date.strftime('%d/%m/%Y') if emp.insurance_end_date else 'مستمر'
            
            emp_data = {
                'employee': emp,
                'code': emp.code,
                'name': emp.name,
                'department_name': emp.department.name if emp.department else 'غير محدد',
                'job_title': emp.job_title or 'غير محدد',
                'insurance_number': emp.insurance_number or 'غير محدد',
                'insurance_policy': emp.insurance_policy or 'employee_only',
                'insurance_start_date': start_date_str,
                'insurance_end_date': end_date_str,
                'insurance_salary': emp.insurance_salary or 0,
                'employee_share_percent': emp.insurance_employee_share or 11.0,
                'company_share_percent': emp.insurance_company_share or 18.75,
                'employee_deduction': insurance_vals.get('employee_deduction', 0),
                'company_cost': insurance_vals.get('company_cost', 0),
                'total_insurance': insurance_vals.get('total_insurance', 0),
            }
            
            report_data.append(emp_data)
            total_employee_deduction += emp_data['employee_deduction']
            total_company_cost += emp_data['company_cost']
            total_insurance += emp_data['total_insurance']
        
        # Sort by department then by name
        report_data.sort(key=lambda x: (x['department_name'], x['name']))
        
        # Get departments for filter
        departments = session.query(Department).all()
        
        # Calculate statistics
        stats = {
            'total_insured': len(report_data),
            'total_employee_deduction': total_employee_deduction,
            'total_company_cost': total_company_cost,
            'total_insurance': total_insurance,
            'average_employee_deduction': total_employee_deduction / len(report_data) if report_data else 0,
            'average_company_cost': total_company_cost / len(report_data) if report_data else 0,
            'average_total': total_insurance / len(report_data) if report_data else 0,
        }
        
        return render_template('reports/insured_employees_detailed.html',
                             report_data=report_data,
                             departments=departments,
                             stats=stats,
                             department_id=department_id,
                             search_query=search_query)
    finally:
        session.close()


@reports_bp.route('/insured_employees/export_excel')
def export_insured_employees_excel():
    """Export detailed insured employees report to Excel"""
    try:
        db = current_app.db
        session = db.get_session()
        
        # Get filter parameters
        department_id = request.args.get('department_id', type=int)
        search_query = request.args.get('search', '').strip()
        
        # Get all insured active employees
        employees = session.query(Employee).filter_by(is_active=True, is_insured=True).all()
        
        # Filter by department if specified
        if department_id:
            employees = [e for e in employees if e.department_id == department_id]
        
        # Apply search filter
        if search_query:
            search_lower = search_query.lower()
            employees = [e for e in employees if 
                        (e.name and search_lower in e.name.lower()) or
                        (e.code and search_lower in e.code.lower()) or
                        (e.insurance_number and search_lower in e.insurance_number.lower())]
        
        # Prepare Excel data
        excel_data = []
        for emp in employees:
            insurance_vals = emp.calculate_insurance_values()
            
            start_date_str = emp.insurance_start_date.strftime('%d/%m/%Y') if emp.insurance_start_date else '-'
            end_date_str = emp.insurance_end_date.strftime('%d/%m/%Y') if emp.insurance_end_date else 'مستمر'
            
            # Policy name mapping
            policy_map = {
                'employee_only': 'خصم الموظف فقط (قانوني)',
                'both_from_employee': 'استقطاع من الموظف (نسبتان)',
                'company_pays_all': 'تحمل الشركة الكاملة'
            }
            policy_name = policy_map.get(emp.insurance_policy, emp.insurance_policy)
            
            excel_data.append({
                'كود الموظف': emp.code,
                'اسم الموظف': emp.name,
                'القسم': emp.department.name if emp.department else '-',
                'الوظيفة': emp.job_title or '-',
                'رقم التأمين': emp.insurance_number or '-',
                'نوع السياسة': policy_name,
                'تاريخ البدء': start_date_str,
                'تاريخ الانتهاء': end_date_str,
                'راتب التأمين': float(emp.insurance_salary or 0),
                'نسبة العامل (%)': float(emp.insurance_employee_share or 11.0),
                'نسبة الشركة (%)': float(emp.insurance_company_share or 18.75),
                'خصم الموظف': float(insurance_vals.get('employee_deduction', 0)),
                'تحمل الشركة': float(insurance_vals.get('company_cost', 0)),
                'الإجمالي': float(insurance_vals.get('total_insurance', 0))
            })
        
        session.close()
        
        if not excel_data:
            flash('لا توجد بيانات للتصدير', 'warning')
            return redirect(url_for('reports.insured_employees'))
        
        # Create Excel file
        df = pd.DataFrame(excel_data)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='المؤمن عليهم')
            apply_professional_style(writer.book.active, df)
        
        output.seek(0)
        filename = f"Insured_Employees_Detailed_{datetime.now().strftime('%Y_%m_%d')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f"حدث خطأ أثناء تصدير ملف Excel: {str(e)}", "danger")
        return redirect(url_for('reports.insured_employees'))


@reports_bp.route('/employees')
def employees():
    """Employees report"""
    db = current_app.db
    all_employees = db.get_all_employees()
    employees = [e for e in all_employees if e.is_active]
    departments = db.get_departments()

    # Attach effective salary for display
    db.attach_effective_salaries(employees)
    
    return render_template('reports/employees.html',
                         employees=employees,
                         departments=departments)

@reports_bp.route('/attendance')
def attendance():
    """Attendance report"""
    db = current_app.db
    
    # Get filter parameters
    date_from_str = request.args.get('date_from')
    date_to_str = request.args.get('date_to')
    
    date_from = None
    date_to = None
    
    if date_from_str:
        parsed = parse_date_compact(date_from_str)
        if parsed:
            date_from = parsed
            
    if date_to_str:
        parsed = parse_date_compact(date_to_str)
        if parsed:
            date_to = parsed

    # Auto-process data for the requested range
    try:
        process_start = date_from if date_from else datetime.now().date()
        process_end = date_to if date_to else datetime.now().date()
        
        # If no dates specific, maybe process last few days or just today?
        # Safe default: process today if nothing specified to ensure real-time
        
        current = process_start
        while current <= process_end:
            db.process_attendance_for_date(current)
            current += timedelta(days=1)
            
    except Exception as e:
        print(f"Error processing attendance: {e}")
    
    attendance_data = db.get_attendance_report(date_from, date_to)
    
    # For UI, pass the formatted string expected DD/MM/YYYY if available
    date_from_ui = format_date_ar(date_from) if date_from else (date_from_str or '')
    date_to_ui = format_date_ar(date_to) if date_to else (date_to_str or '')

    return render_template('reports/attendance.html',
                         data=attendance_data,
                         date_from=date_from_ui,
                         date_to=date_to_ui)

@reports_bp.route('/payroll_sheet')
def payroll_sheet():
    """Department Payroll Sheet (Comprehensive)"""
    db = current_app.db
    session = db.get_session()
    # Get all departments for filter
    all_departments = db.get_departments()
    
    # Filters
    dept_ids = request.args.getlist('department_ids', type=int)
    # Support legacy single ID
    if not dept_ids and request.args.get('dept_id'):
        dept_ids = [request.args.get('dept_id', type=int)]
        
    month = request.args.get('month', type=int, default=datetime.now().month)
    year = request.args.get('year', type=int, default=datetime.now().year)
    
    # Prepare data structure
    # [ {'department': dept_obj, 'employees': [payroll_records], 'total_wages': float, 'count': int}, ... ]
    report_data = []
    
    # Determine which departments to process
    if dept_ids:
        departments_to_process = [d for d in all_departments if d.id in dept_ids]
    else:
        departments_to_process = [] # Default empty or all? Usually empty until search 
        # But if user wants "All", they might send no IDs? Let's assume empty means "Select Departments" state
        # Or if "All" is desired, we can pass flag. For now, empty list = no report.
        if request.args.get('action') == 'view_all':
             departments_to_process = all_departments

    calculator = PayrollCalculator(db)
    
    try:
        for dept in departments_to_process:
            dept_record = {
                'department': dept,
                'employees': [],
                'total_wages': 0.0,
                'count': 0
            }
            
            # Get active employees in this department with only needed fields
            # Faster than accessing dept.employees relationship which loads full objects
            from sqlalchemy.orm import load_only
            active_employees = session.query(Employee).options(
                load_only(
                    Employee.id, Employee.code, Employee.name, Employee.job_title, 
                    Employee.is_active, Employee.department_id, Employee.basic_salary
                )
            ).filter_by(department_id=dept.id, is_active=True).order_by(Employee.code.asc()).all()
            
            for emp in active_employees:
                try:
                    # Calculate payroll
                    salary_info = calculator.calculate_monthly_payroll(emp.id, month, year)
                    # Create employee record
                    employee_data = {
                        'code': emp.code,
                        'name': emp.name,
                        'job_title': emp.job_title,
                        'is_active': emp.is_active
                    }
                    
                    # Update with calculated financial data
                    # salary_info contains: Basic Salary, Net Salary, Total Additions, Total Deductions, etc.
                    employee_data.update(salary_info)
                    
                    # Add specific computed fields for the report if not present
                    if 'Bonuses' not in employee_data:
                         # Fallback if calculator varies
                         employee_data['Bonuses'] = salary_info.get('Incentive', 0) + salary_info.get('Bonuses', 0)
                         
                    dept_record['employees'].append(employee_data)
                    
                    dept_record['total_wages'] += salary_info.get('Net Salary', 0.0)
                    dept_record['count'] += 1
                except Exception as e:
                    print(f"Error calculating payroll for {emp.name}: {e}")
                    
            if dept_record['employees']:
                report_data.append(dept_record)
    finally:
        session.close()
            
    return render_template('reports/payroll_sheet.html',
                         all_departments=all_departments,
                         report_data=report_data,
                         selected_dept_ids=dept_ids,
                         month=month,
                         year=year)


@reports_bp.route('/payroll_signature')
def payroll_signature():
    """Payroll Signature Sheet (Simple)"""
    db = current_app.db
    all_departments = db.get_departments()

    dept_ids = request.args.getlist('department_ids', type=int)
    month = request.args.get('month', type=int, default=datetime.now().month)
    year = request.args.get('year', type=int, default=datetime.now().year)
    
    report_data = []
    
    if dept_ids:
        departments_to_process = [d for d in all_departments if d.id in dept_ids]
    elif request.args.get('action') == 'view_all':
        departments_to_process = all_departments
    else:
        departments_to_process = []

    calculator = PayrollCalculator(db)
    
    for dept in departments_to_process:
        dept_record = {
            'department': dept,
            'employees': [],
            'total_wages': 0.0,
            'count': 0
        }
        
        active_employees = sorted(
            [e for e in dept.employees if e.is_active],
            key=lambda e: (e.code or '')
        )
        
        for emp in active_employees:
            try:
                salary_info = calculator.calculate_monthly_payroll(emp.id, month, year)
                
                emp_data = {
                    'code': emp.code,
                    'name': emp.name,
                    'actual_days': salary_info.get('Actual Days', 0),
                    'Net Salary': salary_info.get('Net Salary', 0)
                }
                
                dept_record['employees'].append(emp_data)
                dept_record['total_wages'] += salary_info.get('Net Salary', 0)
                dept_record['count'] += 1
                
            except Exception as e:
                print(f"Error in signature sheet for {emp.name}: {e}")
        
        if dept_record['employees']:
            report_data.append(dept_record)
            
    return render_template('reports/payroll_signature.html',
                         all_departments=all_departments,
                         report_data=report_data,
                         selected_dept_ids=dept_ids,
                         month=month,
                         year=year,
                         now=datetime.now())


@reports_bp.route('/payroll')
def payroll():
    """Simple Monthly Payroll Report (Flat)"""
    db = current_app.db
    month = request.args.get('month', type=int, default=datetime.now().month)
    year = request.args.get('year', type=int, default=datetime.now().year)

    calculator = PayrollCalculator(db)
    all_employees = db.get_all_employees()
    
    salaries = []
    for emp in all_employees:
        if not emp.is_active:
            continue
            
        try:
            # Use calculate_monthly_payroll to get dict result
            salary_data = calculator.calculate_monthly_payroll(emp.id, month, year)
            
            salaries.append({
                'code': emp.code,
                'name': emp.name,
                'basic_salary': salary_data.get('Gross Salary', 0),
                'total_additions': salary_data.get('Total Additions', 0),
                'total_deductions': salary_data.get('Total Deductions', 0),
                'net_salary': salary_data.get('Net Salary', 0)
            })
        except Exception as e:
            print(f"Skipping {emp.name} in flat payroll: {e}")

    return render_template('reports/payroll.html', 
                         salaries=salaries,
                         month=month, 
                         year=year)


@reports_bp.route('/loans')
@login_required
@permission_required('view_loans_report')
def loans():
    """Loans report"""
    db = current_app.db
    loans = db.get_all_loans()
    # Filter for active employees only
    loans = [l for l in loans if l.employee and l.employee.is_active]
    
    return render_template('reports/loans.html', loans=loans)

@reports_bp.route('/detailed_salary')
def detailed_salary_index():
    """Select employee for detailed report"""
    db = current_app.db
    db = current_app.db
    all_employees = db.get_all_employees()
    employees = [e for e in all_employees if e.is_active]
    return render_template('reports/select_employee.html', employees=employees)

@reports_bp.route('/detailed_salary/<int:emp_id>')
def detailed_salary(emp_id):
    """Generate detailed salary report for employee"""
    session = None
    try:
        db = current_app.db
        month = request.args.get('month', type=int, default=datetime.now().month)
        year = request.args.get('year', type=int, default=datetime.now().year)

        session = db.get_session()
        emp = session.query(Employee).filter_by(id=emp_id).first()
        if not emp:
            flash('الموظف غير موجود', 'danger')
            return redirect(url_for('reports.detailed_salary_index'))
        
        calculator = PayrollCalculator(db)
        report_data = calculator.get_detailed_payroll_report(emp_id, month, year)

        selected_department_ids = request.args.getlist('department_ids', type=int)
        departments = db.get_departments()

        # Navigation list: active employees only (id + name), without payroll recalculation.
        employees_query = session.query(Employee.id, Employee.name).filter(
            Employee.is_active == True
        )
        if selected_department_ids:
            employees_query = employees_query.filter(Employee.department_id.in_(selected_department_ids))
        payroll_employees = employees_query.order_by(Employee.code.asc()).all()
        payroll_employees = [
            {'id': emp_id, 'name': emp_name}
            for emp_id, emp_name in payroll_employees
        ]

        return render_template(
            'reports/detailed_salary.html',
            report=report_data,
            payroll_employees=payroll_employees,
            departments=departments,
            selected_department_ids=selected_department_ids
        )
        
    except Exception as e:
        flash(f"حدث خطأ أثناء إنشاء التقرير: {str(e)}", "danger")
        return redirect(safe_referrer('reports.index'))
    finally:
        if session is not None:
            session.close()


@reports_bp.route('/interactive_detailed')
@login_required
@permission_required('view_interactive_detailed_salary')
def interactive_detailed_index():
    """Index for selecting an employee for interactive report"""
    db = current_app.db
    all_employees = db.get_all_employees()
    employees = [e for e in all_employees if e.is_active]
    return render_template('reports/select_employee_interactive.html', employees=employees)


@reports_bp.route('/interactive_detailed/<int:emp_id>')
@login_required
@permission_required('view_interactive_detailed_salary')
def interactive_detailed_salary(emp_id):
    """Interactive Detailed Salary Report for direct management"""
    session = None
    try:
        db = current_app.db
        month = request.args.get('month', type=int, default=datetime.now().month)
        year = request.args.get('year', type=int, default=datetime.now().year)

        session = db.get_session()
        emp = session.query(Employee).filter_by(id=emp_id).first()
        if not emp:
            flash('الموظف غير موجود', 'danger')
            return redirect(url_for('reports.detailed_salary_index'))
        
        calculator = PayrollCalculator(db)
        report_data = calculator.get_detailed_payroll_report(emp_id, month, year)

        selected_department_ids = request.args.getlist('department_ids', type=int)
        departments = db.get_departments()

        # Navigation list
        employees_query = session.query(Employee.id, Employee.name).filter(Employee.is_active == True)
        if selected_department_ids:
            employees_query = employees_query.filter(Employee.department_id.in_(selected_department_ids))
        payroll_employees = [{'id': eid, 'name': ename} for eid, ename in employees_query.order_by(Employee.code.asc()).all()]

        return render_template(
            'reports/interactive_detailed_salary.html',
            report=report_data,
            payroll_employees=payroll_employees,
            departments=departments,
            selected_department_ids=selected_department_ids
        )
    except Exception as e:
        flash(f"حدث خطأ: {str(e)}", "danger")
        return redirect(url_for('reports.index'))
    finally:
        if session is not None:
            session.close()


@reports_bp.route('/get_employees_by_departments', methods=['GET'])
def get_employees_by_departments():
    """Return active employees (id + name) filtered by selected departments."""
    db = current_app.db
    session = db.get_session()
    try:
        department_ids = request.args.getlist('departments_ids[]', type=int)
        if not department_ids:
            department_ids = request.args.getlist('departments_ids', type=int)
        if not department_ids:
            department_ids = request.args.getlist('department_ids', type=int)

        query = session.query(Employee.id, Employee.name).filter(Employee.is_active == True)
        if department_ids:
            query = query.filter(Employee.department_id.in_(department_ids))

        employees = query.order_by(Employee.code.asc()).all()
        return jsonify([{'id': emp_id, 'name': emp_name} for emp_id, emp_name in employees])
    except Exception:
        current_app.logger.exception('Failed to load employees by departments')
        return jsonify([]), 500
    finally:
        session.close()

@reports_bp.route('/employee_history/<employee_code>')
def employee_history(employee_code):
    """
    تقرير تتبع التعديلات على بيانات الموظف
    يعرض جميع التغييرات التي تمت على البيانات الأساسية للموظف
    """
    db = current_app.db

    # الخطوة 1: جلب الموظف
    employee = db.get_employee_by_code(employee_code)
    
    if not employee:
        flash(f"لم يتم العثور على موظف بالكود {employee_code}", "danger")
        return redirect(url_for('reports.index'))

    # الخطوة 2: جلب سجلات التتبع للموظف
    history_records = db.get_audit_logs_by_employee(employee_code)

    # الخطوة 3: عرض القالب
    return render_template('reports/audit_report.html',
                           employee=employee,
                           history_records=history_records)


@reports_bp.route('/audit_trail')
def audit_trail():
    """
    تقرير تتبع شامل لجميع التعديلات على بيانات الموظفين
    يعرض آخر التعديلات مع إمكانية البحث والتصفية
    """
    db = current_app.db
    
    # الحصول على معاملات البحث
    employee_code = request.args.get('employee_code', '').strip()
    field_name = request.args.get('field_name', '').strip()
    limit = request.args.get('limit', type=int, default=100)
    
    # الحصول على قائمة الموظفين لخيار البحث
    employees = db.get_all_employees(only_active=True)
    
    # الحصول على التقرير
    if employee_code and field_name:
        # البحث حسب الموظف والحقل
        logs = db.get_audit_log_history(employee_code, field_name)
    elif employee_code:
        # البحث حسب الموظف فقط
        logs = db.get_audit_logs_by_employee(employee_code, limit=limit)
    elif field_name:
        # البحث حسب الحقل فقط
        logs = db.get_audit_logs_by_field(field_name, limit=limit)
    else:
        # آخر التعديلات
        logs = db.get_audit_logs_recent(limit=limit)
    
    # إحصائيات
    if employee_code:
        summary = db.get_audit_log_summary(employee_code)
    else:
        summary = None
    
    return render_template('reports/audit_trail.html',
                           logs=logs,
                           employees=employees,
                           summary=summary,
                           selected_employee=employee_code,
                           selected_field=field_name)


@reports_bp.route('/audit_export')
def audit_export():
    """
    تصدير تقرير التتبع إلى ملف CSV
    """
    db = current_app.db
    
    try:
        filename = f"audit_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        filepath = os.path.join('/tmp', filename)  # الملف المؤقت
        
        # تصدير السجلات
        success = db.export_audit_logs_csv(filepath)
        
        if success:
            # إرسال الملف للتحميل
            from flask import send_file
            return send_file(filepath, as_attachment=True, download_name=filename)
        else:
            flash("فشل تصدير التقرير", "danger")
            return redirect(url_for('reports.audit_trail'))
            
    except Exception as e:
        flash(f"خطأ في التصدير: {str(e)}", "danger")
        return redirect(url_for('reports.audit_trail'))


@reports_bp.route('/permanent_loans')
@login_required
@permission_required('view_loans_report')
def permanent_loans():
    """Detailed Permanent Loans report"""
    db = current_app.db
    dept_id = request.args.get('dept_id', type=int)
    
    all_loans = db.get_all_loans()
    # Filter: permanent and not paid off
    target_loans = [l for l in all_loans if l.type == 'permanent' and not l.is_paid_off]
    
    if dept_id:
        target_loans = [l for l in target_loans if l.employee and l.employee.department_id == dept_id and l.employee.is_active]
    else:
        # Filter for active employees even if no department selected
        target_loans = [l for l in target_loans if l.employee and l.employee.is_active]
        
    report_data = []
    for loan in target_loans:
        deduction_months = []
        excluded = []
        if loan.excluded_months:
            for tok in loan.excluded_months.split(','):
                t = tok.strip()
                if not t:
                    continue
                try:
                    excluded.append(int(t))
                except ValueError:
                    try:
                        f = float(t)
                        if f.is_integer():
                            excluded.append(int(f))
                    except ValueError:
                        continue
        
        if loan.date:
            import calendar
            # Always start schedule from the loan issuance date for verification
            current_date = loan.date
            
            # Show the FULL original schedule (all installments)
            total_count = loan.installments_count
            installments_accounted = 0
            
            # Loop for up to 10 years
            for _ in range(120):
                if installments_accounted >= total_count:
                    break
                
                # Current month info
                year = current_date.year
                month = current_date.month
                
                if month not in excluded:
                    deduction_months.append(f"{month:02d}/{year}")
                    installments_accounted += 1
                
                # Move to next month
                month += 1
                if month > 12:
                    month = 1
                    year += 1
                
                # Get valid day for next month
                last_day = calendar.monthrange(year, month)[1]
                day = min(loan.date.day, last_day) # Keep original day if possible
                current_date = date(year, month, day)
        
        report_data.append({
            'loan': loan,
            'deduction_months': deduction_months,
            'excluded_months_list': excluded,
            'current_balance': loan.auto_remaining_balance  # New dynamic field
        })
        
    departments = db.get_departments()
    return render_template('reports/permanent_loans.html', 
                         data=report_data, 
                         departments=departments,
                         selected_dept=dept_id,
                         now=datetime.now())


@reports_bp.route('/detailed_salary/<int:emp_id>/export_excel')
def export_detailed_salary_excel(emp_id):
    """Export detailed salary report to Excel (Single Sheet + Professional Styling)"""
    try:
        db = current_app.db
        month = request.args.get('month', type=int, default=datetime.now().month)
        year = request.args.get('year', type=int, default=datetime.now().year)
        
        calculator = PayrollCalculator(db)
        report = calculator.get_detailed_payroll_report(emp_id, month, year)
        
        # Prepare Summary Section
        summary = report['summary']
        summary_rows = [
            ['اسم الموظف', report['employee_name'], 'كود الموظف', report['employee_code']],
            ['الشهر', f"{month} - {year}", 'القسم', report['department_name']],
            ['', '', '', ''],
            ['بند المستحقات', 'القيمة', 'بند الاستقطاعات', 'القيمة'],
            ['الراتب الأساسي', float(report['basic_salary']), 'خصم التأخير', float(summary['late_deduction'])],
            ['الحوافز', float(summary['incentive']), 'خصم انصراف مبكر', float(summary['early_deduction'])],
            ['حافز الانتظام', float(summary['regularity_incentive']), 'جزاءات الغياب', float(summary['absence_penalty'])],
            ['المكافآت', float(summary['bonuses']), 'خصم التصاريح', float(summary['permissions_deduction'])],
            ['قيمة الإضافي', float(summary['overtime_value']), 'أقساط السلف', float(summary['loans_deduction'])],
            ['بدل انتقال', float(summary['transport']), 'جزاءات إدارية', float(summary['admin_penalties'])],
            ['أيام الحضور', float(summary['attendance_days']), 'تأمينات اجتماعية', float(summary['insurance'])],
            ['إجمالي المستحقات', float(summary['total_additions']), 'إجمالي الاستقطاعات', float(summary['total_deductions'])],
            ['', '', 'صافي الراتب', float(summary['net_salary'])]
        ]
        
        # Prepare Daily Details Data
        daily_details = []
        for d in report['daily_details']:
            daily_details.append({
                'التاريخ': d['date'],
                'حالة اليوم': d['status'],
                'دخول': d['check_in'].strftime('%H:%M') if d['check_in'] else '-',
                'خروج': d['check_out'].strftime('%H:%M') if d['check_out'] else '-',
                'تأخير(د)': int(d['late_minutes']),
                'خصم تأخير': float(d['late_deduction']),
                'انصراف مبكر': float(d['early_deduction']),
                'إضافي(س)': float(d['overtime_hours']),
                'قيمة إضافي': float(d['overtime_value']),
                'تصاريح(س)': float(d['permission_hours']),
                'خصم تصاريح': float(d['permission_deduction'])
            })
            
        output = io.BytesIO()
        # Use pandas just for easier data prep, but openpyxl for everything else
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # We'll write data manually using openpyxl engine
            workbook = writer.book
            sheet = workbook.create_sheet('التقرير التفصيلي', 0)
            sheet.sheet_view.rightToLeft = True
            
            # Styles
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            border_side = Side(style='thin', color='000000')
            full_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
            center_align = Alignment(horizontal='center', vertical='center')
            
            # 1. Write Summary Header
            sheet.merge_cells('A1:D1')
            sheet['A1'] = f"تقرير مفردات المرتب - {report['employee_name']}"
            sheet['A1'].font = Font(size=14, bold=True)
            sheet['A1'].alignment = center_align
            
            # 2. Write Summary Rows
            for r_idx, row in enumerate(summary_rows, start=2):
                for c_idx, val in enumerate(row, start=1):
                    cell = sheet.cell(row=r_idx, column=c_idx, value=val)
                    cell.alignment = center_align
                    if r_idx == 5: # Financial Table Header
                         cell.fill = header_fill
                         cell.font = header_font
                    if r_idx >= 5 and r_idx <= 14 and c_idx <= 4:
                         cell.border = full_border

            # 3. Write Daily Details Header
            start_daily_row = 17
            sheet.merge_cells(f'A{start_daily_row}:K{start_daily_row}')
            sheet[f'A{start_daily_row}'] = "تفاصيل الحضور والانصراف اليومية"
            sheet[f'A{start_daily_row}'].font = Font(size=12, bold=True)
            sheet[f'A{start_daily_row}'].alignment = center_align
            
            # 4. Write Details Table
            df_daily = pd.DataFrame(daily_details)
            for c_idx, col in enumerate(df_daily.columns, start=1):
                cell = sheet.cell(row=start_daily_row + 1, column=c_idx, value=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = full_border
                
            for r_idx, row in enumerate(df_daily.values, start=start_daily_row + 2):
                for c_idx, val in enumerate(row, start=1):
                    cell = sheet.cell(row=r_idx, column=c_idx, value=val)
                    cell.alignment = center_align
                    cell.border = full_border

            # Auto-adjust column width
            for col in sheet.columns:
                max_length = 0
                # Get column letter safely (handle merged cells)
                try:
                    column = col[0].column_letter
                except (AttributeError, IndexError):
                    # For merged cells, try to find the column from the first non-merged cell
                    column = None
                    for cell in col:
                        if hasattr(cell, 'column_letter'):
                            column = cell.column_letter
                            break
                    if not column:
                        continue
                
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except: pass
                sheet.column_dimensions[column].width = max_length + 2

        output.seek(0)
        filename = f"Detailed_Salary_{report['employee_code']}_{month}_{year}.xlsx"
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=filename)
    except Exception as e:
        flash(f"حدث خطأ أثناء تصدير ملف Excel: {str(e)}", "danger")
        return redirect(safe_referrer('reports.index'))

@reports_bp.route('/payroll_sheet/export_excel')
def export_payroll_sheet_excel():
    """Export Department Payroll Sheet to Excel (Clean Numeric Content + Professional Style)"""
    try:
        db = current_app.db
        dept_ids = request.args.getlist('department_ids', type=int)
        if not dept_ids and request.args.get('dept_id'):
            dept_ids = [request.args.get('dept_id', type=int)]
            
        month = request.args.get('month', type=int, default=datetime.now().month)
        year = request.args.get('year', type=int, default=datetime.now().year)
        
        all_departments = db.get_departments()
        if dept_ids:
            departments_to_process = [d for d in all_departments if d.id in dept_ids]
        else:
            departments_to_process = all_departments if request.args.get('action') == 'view_all' else []

        if not departments_to_process:
            flash("يرجى اختيار الأقسام المراد تصديرها أولاً", "warning")
            return redirect(safe_referrer('reports.payroll_sheet'))

        calculator = PayrollCalculator(db)
        excel_data = []

        for dept in departments_to_process:
            active_employees = [e for e in dept.employees if e.is_active]
            for emp in active_employees:
                try:
                    salary_info = calculator.calculate_monthly_payroll(emp.id, month, year)
                    
                    # Create a flat record with numeric values
                    record = {
                        'القسم': dept.name,
                        'كود الموظف': emp.code,
                        'الاسم': emp.name,
                        'الوظيفة': emp.job_title,
                        'الراتب الأساسي': float(salary_info.get('Basic Salary', 0)),
                        'أيام الحضور': float(salary_info.get('Attendance Days', 0)),
                        'الراتب الإجمالي': float(salary_info.get('Gross Salary', 0)),
                        'حوافز': float(salary_info.get('Incentive', 0) + salary_info.get('Regularity Incentive', 0) + salary_info.get('Bonuses', 0)),
                        'إضافي': float(salary_info.get('OT Value', 0)),
                        'بدلات': float(salary_info.get('Transport Allowance', 0)),
                        'إجمالي المستحقات': float(salary_info.get('Total Additions', 0)),
                        'تأمينات': float(salary_info.get('Insurance', 0)),
                        'تأخيرات': float(salary_info.get('Lateness Deduction', 0)),
                        'انصراف مبكر': float(salary_info.get('Early Deduction', 0)),
                        'جزاءات': float(salary_info.get('Absence Penalty Deduction', 0) + salary_info.get('Admin Penalties', 0)),
                        'سلف': float(salary_info.get('Loan Deduction', 0)),
                        'إجمالي الاستقطاعات': float(salary_info.get('Total Deductions', 0)),
                        'صافي الراتب': float(salary_info.get('Net Salary', 0))
                    }
                    excel_data.append(record)
                except:
                    continue

        if not excel_data:
            flash("لا توجد بيانات لتصديرها", "warning")
            return redirect(safe_referrer('reports.payroll_sheet'))

        # Create Excel in memory
        df = pd.DataFrame(excel_data)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Payroll Sheet')
            apply_professional_style(writer.book.active, df)

        output.seek(0)
        filename = f"Payroll_Sheet_{month}_{year}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f"حدث خطأ أثناء تصدير ملف Excel: {str(e)}", "danger")
        return redirect(safe_referrer('reports.payroll_sheet'))

@reports_bp.route('/insurance_costs/export_excel')
def export_insurance_costs_excel():
    """Export Insurance Costs Report to Excel (Clean Numeric Content + Professional Style)"""
    try:
        db = current_app.db
        session = db.get_session()
        
        departments = session.query(Department).all()
        employees = session.query(Employee).filter_by(is_active=True, is_insured=True).all()
        
        excel_data = []
        for dept in departments:
            dept_emps = [e for e in employees if e.department_id == dept.id]
            if not dept_emps:
                continue
                
            stats = {
                'القسم': dept.name,
                'عدد المؤمن عليهم': len(dept_emps),
                'خصم الموظفين': 0.0,
                'تحمل الشركة': 0.0,
                'الإجمالي الكلي': 0.0
            }
            
            for emp in dept_emps:
                vals = emp.calculate_insurance_values()
                stats['خصم الموظفين'] += float(vals['employee_deduction'])
                stats['تحمل الشركة'] += float(vals['company_cost'])
                stats['الإجمالي الكلي'] += float(vals['total_insurance'])
            
            excel_data.append(stats)
            
        session.close()

        if not excel_data:
            flash("لا توجد بيانات تأمينية لتصديرها", "warning")
            return redirect(url_for('reports.insurance_costs'))

        # Create Excel
        df = pd.DataFrame(excel_data)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Insurance Analysis')
            apply_professional_style(writer.book.active, df)

        output.seek(0)
        filename = f"Insurance_Costs_Report_{datetime.now().strftime('%Y_%m_%d')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f"حدث خطأ أثناء تصدير ملف Excel: {str(e)}", "danger")
        return redirect(url_for('reports.insurance_costs'))


# ===== Salary History Report (تقرير السجل التاريخي للرواتب) =====

@reports_bp.route('/salary_history')
def salary_history():
    """تقرير السجل التاريخي لتعديلات الرواتب"""
    db = current_app.db
    
    # Get filter parameters
    employee_id = request.args.get('employee_id', type=int)
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    department_id = request.args.get('department_id', type=int)
    
    # Parse dates
    date_from = None
    date_to = None
    if from_date:
        try:
            date_from = parse_date_compact(from_date) if from_date else None
        except:
            date_from = None
    
    if to_date:
        try:
            date_to = parse_date_compact(to_date) if to_date else None
        except:
            date_to = None
    
    # Get salary history records
    history_records = db.get_salary_history_report(
        employee_id=employee_id,
        from_date=date_from,
        to_date=date_to
    )
    
    # Get employees and departments for filters
    employees = db.get_all_employees()
    departments = db.get_departments()
    
    # Calculate statistics
    total_increases = sum(h.salary_change for h in history_records if h.salary_change > 0)
    total_decreases = sum(h.salary_change for h in history_records if h.salary_change < 0)
    change_count = len(history_records)
    
    return render_template('reports/salary_history.html',
        history_records=history_records,
        employees=employees,
        departments=departments,
        employee_id=employee_id,
        from_date=from_date,
        to_date=to_date,
        department_id=department_id,
        total_increases=total_increases,
        total_decreases=total_decreases,
        change_count=change_count
    )


@reports_bp.route('/salary_history/<int:emp_id>')
def salary_history_employee(emp_id):
    """تقرير السجل التاريخي لموظف محدد"""
    db = current_app.db
    
    employee = db.get_employee_by_id(emp_id)
    if not employee:
        flash('الموظف غير موجود', 'danger')
        return redirect(url_for('reports.salary_history'))

    # Get salary history
    history = db.get_employee_salary_history(emp_id)

    # Attach effective salary for display
    db.attach_effective_salaries([employee])
    
    # Calculate statistics
    total_increase = sum(h.salary_change for h in history if h.salary_change > 0)
    total_decrease = sum(h.salary_change for h in history if h.salary_change < 0)
    average_salary = sum(h.new_salary for h in history) / len(history) if history else (employee.effective_salary if getattr(employee, 'effective_salary', None) is not None else employee.basic_salary)
    
    return render_template('reports/salary_history_employee.html',
        employee=employee,
        history=history,
        total_increase=total_increase,
        total_decrease=total_decrease,
        average_salary=average_salary
    )


@reports_bp.route('/salary_history/export')
def export_salary_history_excel():
    """تصدير تقرير السجل التاريخي للرواتب إلى Excel"""
    db = current_app.db
    
    # Get filter parameters
    employee_id = request.args.get('employee_id', type=int)
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    
    # Parse dates
    date_from = None
    date_to = None
    if from_date:
        try:
            date_from = parse_date_compact(from_date)
        except:
            date_from = None
    
    if to_date:
        try:
            date_to = parse_date_compact(to_date)
        except:
            date_to = None
    
    # Get salary history
    history_records = db.get_salary_history_report(
        employee_id=employee_id,
        from_date=date_from,
        to_date=date_to
    )
    
    # Prepare data for Excel
    data = []
    for record in history_records:
        data.append({
            'رقم الموظف': record.employee.code if record.employee else '',
            'اسم الموظف': record.employee.name if record.employee else '',
            'تاريخ التعديل': format_date_ar(record.change_date.date()) if record.change_date else '',
            'الراتب السابق': f"{record.old_salary:,.2f}",
            'الراتب الجديد': f"{record.new_salary:,.2f}",
            'التغيير': f"{record.salary_change:,.2f}",
            'نوع التغيير': record.change_type,
            'السبب': record.reason or '',
            'ملاحظات': record.notes or '',
            'المعدل بواسطة': record.modified_by or ''
        })
    
    if not data:
        flash('لا توجد بيانات لتصديرها', 'warning')
        return redirect(url_for('reports.salary_history'))
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Create Excel file
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='تقرير الرواتب')
        apply_professional_style(writer.book.active, df)
    
    output.seek(0)
    filename = f"Salary_History_{datetime.now().strftime('%Y_%m_%d')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@reports_bp.route('/documents_status')
def documents_status():
    """Report showing documents provided vs missing per employee"""
    db = current_app.db
    department_id = request.args.get('department_id', type=int)
    search_query = request.args.get('search', '').strip()
    
    # Additional filters from UI
    def _truthy(val):
        if val is None:
            return False
        v = str(val).lower()
        return v in ('1', 'true', 'yes', 'on')

    only_missing = _truthy(request.args.get('only_missing'))
    only_expired = _truthy(request.args.get('only_expired'))
    include_optional = not _truthy(request.args.get('exclude_optional'))

    report = db.get_documents_status_all_employees(
        department_id=department_id,
        only_missing=only_missing,
        only_expired=only_expired,
        include_optional=include_optional
    )

    # Apply search filter
    if search_query:
        search_lower = search_query.lower()
        report = [r for r in report if 
                  (r['employee'].name and search_lower in r['employee'].name.lower()) or
                  (r['employee'].code and search_lower in r['employee'].code.lower())]

    # Calculate statistics and enrich data
    stats = {
        'total_employees': len(report),
        'complete_count': 0,
        'incomplete_count': 0,
        'expired_count': 0,
        'total_required': 0,
        'total_provided': 0
    }
    
    enriched_report = []
    for r in report:
        # Count provided and missing
        provided_count = len(r.get('provided', []))
        missing_count = len(r.get('missing', []))
        required_count = provided_count + missing_count
        
        # Calculate completion percentage
        completion_pct = (provided_count / required_count * 100) if required_count > 0 else 0
        
        # Check for expired documents
        has_expired = any(p.get('is_expired', False) for p in r.get('provided', []))
        
        # Determine row status
        if has_expired:
            row_status = 'expired'
            stats['expired_count'] += 1
        elif completion_pct == 100:
            row_status = 'complete'
            stats['complete_count'] += 1
        else:
            row_status = 'incomplete'
            stats['incomplete_count'] += 1
        
        # Accumulate totals
        stats['total_required'] += required_count
        stats['total_provided'] += provided_count
        
        # Enrich row data
        enriched_row = r.copy()
        enriched_row['provided_count'] = provided_count
        enriched_row['missing_count'] = missing_count
        enriched_row['required_count'] = required_count
        enriched_row['completion_pct'] = round(completion_pct, 1)
        enriched_row['row_status'] = row_status
        enriched_row['has_expired'] = has_expired
        
        enriched_report.append(enriched_row)
    
    # Calculate overall completion percentage
    stats['overall_completion'] = round(
        (stats['total_provided'] / stats['total_required'] * 100) 
        if stats['total_required'] > 0 else 0, 
        1
    )

    departments = db.get_departments()

    return render_template('reports/documents_status.html', 
        report=enriched_report, 
        departments=departments,
        stats=stats,
        department_id=department_id, 
        only_missing=only_missing, 
        only_expired=only_expired, 
        include_optional=include_optional,
        search_query=search_query)


@reports_bp.route('/documents_status/export')
def documents_status_export():
    db = current_app.db
    department_id = request.args.get('department_id', type=int)
    search_query = request.args.get('search', '').strip()
    
    def _truthy(val):
        if val is None:
            return False
        v = str(val).lower()
        return v in ('1', 'true', 'yes', 'on')

    only_missing = _truthy(request.args.get('only_missing'))
    only_expired = _truthy(request.args.get('only_expired'))
    include_optional = not _truthy(request.args.get('exclude_optional'))

    report = db.get_documents_status_all_employees(
        department_id=department_id,
        only_missing=only_missing,
        only_expired=only_expired,
        include_optional=include_optional
    )
    
    # Apply search filter
    if search_query:
        search_lower = search_query.lower()
        report = [r for r in report if 
                  (r['employee'].name and search_lower in r['employee'].name.lower()) or
                  (r['employee'].code and search_lower in r['employee'].code.lower())]

    # Build rows with enhanced data
    rows = []
    for r in report:
        emp = r['employee']
        
        # Calculate counts
        provided_count = len(r.get('provided', []))
        missing_count = len(r.get('missing', []))
        required_count = provided_count + missing_count
        completion_pct = (provided_count / required_count * 100) if required_count > 0 else 0
        
        # Build text lists
        provided_items = []
        for p in r.get('provided', []):
            pname = p.get('type_name')
            if p.get('expiry_date'):
                expiry = p.get('expiry_date')
                expired_flag = ' (منتهي)' if p.get('is_expired') else ''
                provided_items.append(f"{pname} - {expiry}{expired_flag}")
            else:
                provided_items.append(pname)

        missing_items = []
        for m in r.get('missing', []):
            mname = m.get('name') if isinstance(m, dict) else m
            if isinstance(m, dict) and (not m.get('is_required', True)):
                mname = f"{mname} (اختياري)"
            missing_items.append(mname)

        provided_names = ', '.join(provided_items) if provided_items else ''
        missing_names = ', '.join(missing_items) if missing_items else ''
        
        # Determine status
        has_expired = any(p.get('is_expired', False) for p in r.get('provided', []))
        if has_expired:
            status = 'منتهي الصلاحية'
        elif completion_pct == 100:
            status = 'مكتمل'
        else:
            status = 'ناقص'

        rows.append({
            'كود الموظف': emp.code,
            'اسم الموظف': emp.name,
            'القسم': emp.department.name if emp.department else '',
            'عدد المستندات المقدمة': int(provided_count),
            'عدد المستندات الناقصة': int(missing_count),
            'الإجمالي المطلوب': int(required_count),
            'نسبة الاكتمال (%)': float(round(completion_pct, 1)),
            'الحالة': status,
            'تفاصيل المقدمة': provided_names,
            'تفاصيل الناقصة': missing_names
        })

    if not rows:
        flash('لا توجد بيانات للتصدير', 'warning')
        return redirect(url_for('reports.documents_status'))

    df = pd.DataFrame(rows)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Documents Status')
        apply_professional_style(writer.book.active, df)

    output.seek(0)
    filename = f"Documents_Status_{datetime.now().strftime('%Y_%m_%d')}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)

